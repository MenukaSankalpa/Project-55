import os
import pandas as pd
from fuzzywuzzy import process
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


df = pd.read_excel("database.xlsx")

search_term = input("Enter description keyword to search: ")


descriptions = df['Description'].astype(str).tolist()


matches = process.extract(search_term, descriptions, limit=5)

print("\nTop Matches:")
print(f"{'No.':<4} {'Inventory ID':<15} {'Description':<40} {'Type':<20} {'Item Class':<20} {'Score':<6}")

for i, (desc, score) in enumerate(matches):
    row = df[df['Description'] == desc].iloc[0]
    print(f"{i+1:<4} {str(row['Inventory ID']):<15} {str(row['Description']):<40} "
        f"{str(row.get('Type', '')):<20} {str(row.get('Item Class', '')):<20} {score:<6}")

selection_input = input("Select match numbers to export (e.g. 1 or 1,3): ")

selected_indices = []
for part in selection_input.split(','):
    part = part.strip()
    if part.isdigit():
        idx = int(part) - 1
        if 0 <= idx < len(matches):
            selected_indices.append(idx)

if not selected_indices:
    print("No valid selections made. Exiting.")
    exit()

# Gather selected rows
selected_rows = pd.concat([
    df[df['Description'] == matches[idx][0]]
    for idx in selected_indices
])

output_file = "selected_output.xlsx"
sheet_name = 'Selected Items'


if os.path.exists(output_file):
    existing_df = pd.read_excel(output_file, sheet_name=sheet_name)

    combined_df = pd.concat([existing_df, selected_rows]).drop_duplicates(subset=['Inventory ID', 'Description']).reset_index(drop=True)
else:
    combined_df = selected_rows


with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    combined_df.to_excel(writer, index=False, sheet_name=sheet_name)


wb = load_workbook(output_file)
ws = wb[sheet_name]

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2

wb.save(output_file)

print(f"\n Record(s) exported to {output_file}. Total rows now: {len(combined_df)}")
